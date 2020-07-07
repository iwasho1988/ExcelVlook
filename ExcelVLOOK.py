#!/usr/bin/env python3
# coding: utf-8

import openpyxl as opx
import json
import os
import sys
import traceback

ANum = "A"
ACol = "B"
ASiz = "C"
AAmount = "D"
ALast = "E"

BNum = "A"
BCol = "B"
BSiz = "C"
BAmount = "D"
BLast = "E"

def get_file_paths(file_name):
    SEPATATER = "/"  # for Linux
    ENCODING = "utf-8" # for Linux
    if os.name == "nt":
        SEPATATER = "\\"  # for Windows
        ENCODING = "Shift-JIS"  # for Windows

    REPLACE_TXT = os.path.dirname(os.path.abspath(__file__)) + SEPATATER + file_name
    C_PATH = os.path.dirname(os.path.abspath(__file__)) + SEPATATER
    return REPLACE_TXT,ENCODING,C_PATH

def read_text(REPLACE_TXT,ENCODING): 
    search_list_dict = {}
    with open(REPLACE_TXT,mode="r",encoding=ENCODING) as f:
        for rows in f.read().split():
            row = rows.split(":")
            bookname = row[0]
            origin_sheet = row[1]
            read_sheet = row[2]
            print("origin_sheet:",str(origin_sheet))
            print("read_sheet:",str(read_sheet))
    return bookname,origin_sheet,read_sheet

def load_available(ws,Num):
    available = 1
    for cell in ws[Num]:
        if cell.value:
            available += 1
        else:
            break
    
    return available

def insert_combine_col(ws,Num,Col,Siz,Last,available):
    ins_col = opx.utils.column_index_from_string(Last)
    ws.insert_cols(ins_col)

    for row in range(available):
        if row <= 1:
            continue
        Num_cell = Num + str(row)
        Col_cell = Col + str(row)
        Siz_cell = Siz + str(row)
        Combine_cell =  ws.cell(row=row, column=ins_col)
        val = "={}&{}&{}".format(Num_cell,Col_cell,Siz_cell)
        Combine_cell.value = val
    return 0


def insert_amount_col(ws,available):
    combine_col = opx.utils.column_index_from_string(BLast)
    amount_col = combine_col + 1
    ws.insert_cols(amount_col)

    for row in range(available):
        if row <= 1:
            continue
        Amount_cell = ws.cell(row=row, column=amount_col)
        read_combine_cell = BLast + str(available)
        org_amount_cell = BAmount + str(row)
        val = "={}".format(org_amount_cell)
        Amount_cell.value = val
    return opx.utils.get_column_letter(amount_col)


def insert_vlook_col(ws,origin_available,read_available,read_sheet,amount_coll):
    combine_col = opx.utils.column_index_from_string(ALast)
    vlook_col = combine_col + 1
    ws.insert_cols(vlook_col)
    last_amount = amount_coll + str(read_available)

    for row in range(origin_available):
        if row <= 1:
            continue
        combine_cell = ALast + str(row)
        read_combine_cell = BLast + str(1)

        Vlook_cell =  ws.cell(row=row, column=vlook_col)
        val = "=VLOOKUP({},{}!{}:{},2,False)".format(combine_cell,read_sheet,read_combine_cell,last_amount)
        Vlook_cell.value = val
    return 0


def main():
    try:
        #args = sys.argv
        #EXCEL_FILE = args[1]
        #EXCEL_SHEET = args[2]
        REPLACE_TXT,ENCODING,C_PATH = get_file_paths("search.txt")
        bookname,origin_sheet,read_sheet = read_text(REPLACE_TXT,ENCODING)
        book = C_PATH + bookname
        print("File path:",book)
        wb = opx.load_workbook(book)

        ws1 = wb[origin_sheet]
        origin_available = load_available(ws1,ANum)

        ws2 = wb[read_sheet]
        read_available = load_available(ws2,BNum)
        print(origin_available,read_available)

        insert_combine_col(ws1,ANum,ACol,ASiz,ALast,origin_available)
        insert_combine_col(ws2,BNum,BCol,BSiz,BLast,read_available)

        amount_coll = insert_amount_col(ws2,read_available)

        insert_vlook_col(ws1,origin_available,read_available,read_sheet,amount_coll)
        wb.save(book)

        return "success!!"
    except Exception as e:
        return str(traceback.format_exc())

if __name__ == '__main__':
    print(main())
